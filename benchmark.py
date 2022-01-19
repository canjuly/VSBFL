# -*- coding: utf-8 -*-
import os
import sys
import util
import time
import VFL
import Parse_ast
import Coverage
import openpyxl
import numpy as np
import SBFL_Formular as SF
import Variable_sus as vs

data_path = r'../ITSP-data'
res_file = ''
res_file_temp = os.path.join(r'C:\Users\ShenJitao\Desktop\cppsnooper\instrumentation\result', 'sbfl(%s)_sus.xlsx')  # 请先自己创建这个文件
now_fm = ''

def find_pair_by_tag(dir_path):
    '''
    根据数据集寻找每份错误代码对应的正确代码
    '''
    pair_info = {}
    tag_files = os.listdir(dir_path)
    for i in tag_files:
        tag_file = os.path.join(dir_path, i)
        lines = util.read_file(tag_file)
        wa_file = lines[0].replace('\n', '')
        ac_file = lines[1].replace('\n', '')
        pair_info[wa_file] = ac_file
    # for file in pair_info:
    #     print(file, pair_info[file])
    return pair_info

def find_pair_by_res(file_path):
    '''
    根据匹配结果寻找每份错误代码对应的正确代码
    '''
    pair_info = {}
    if file_path.split('.')[-1] != 'xlsx' and file_path.split('.')[-1] != 'xls':
        print('not an execl file')
        return
    wb = openpyxl.load_workbook(file_path)
    for i, sheet in enumerate(wb):
        if i == 0:
            continue
        # print(sheet.title)
        buggy_name = ''
        for j, row in enumerate(sheet.rows):
            if j == 0:
                continue
            if row[0].value != None:
                buggy_name = row[0].value
            if row[1].value.find('[') >= 0:
                ac_name_list = eval(row[1].value)
                pair_info[buggy_name] = ac_name_list[0]
    #     break
    # print(pair_info)
    return pair_info

def cal_N_tuple(passed_test_num, failed_test_num, lines_passed,  lines_failed):
    '''
    计算各行的怀疑度值
    这里可以切换相似系数（我不知道我相似系数那几个公式对不对。。）
    '''
    N_tuple = []
    line_num = len(lines_passed)  #其实比实际行数多一行，因为有个第0行
    for i in range(1, line_num):
        Ncf = lines_failed[i]
        Nuf = failed_test_num - lines_failed[i]
        Ncp = lines_passed[i]
        Nup = passed_test_num - lines_passed[i]
        sus = 0
        if now_fm == 'Jaccard':
            sus = SF.cal_jaccard(Ncf, Nuf, Ncp, Nup)
        if now_fm == 'Tarantula':
            sus = SF.cal_turantula(Ncf, Nuf, Ncp, Nup)
        if now_fm == 'Dstar':
            sus = SF.cal_dstar(Ncf, Nuf, Ncp, Nup)
        if now_fm == 'Ochiai':
            sus = SF.cal_ochiai(Ncf, Nuf, Ncp, Nup)
        if now_fm == 'Op2':
            sus = SF.cal_op2(Ncf, Nuf, Ncp, Nup)
        # Naish = SF.cal_naish(Ncf, Nuf, Ncp, Nup)
        # GP08 = SF.cal_GP08(Ncf, Nuf, Ncp, Nup)
        # GP10 = SF.cal_GP10(Ncf, Nuf, Ncp, Nup)
        # GP11 = SF.cal_GP11(Ncf, Nuf, Ncp, Nup)
        # GP13 = SF.cal_GP13(Ncf, Nuf, Ncp, Nup)
        # GP20 = SF.cal_GP20(Ncf, Nuf, Ncp, Nup)
        # GP26 = SF.cal_GP26(Ncf, Nuf, Ncp, Nup)
        # print(i, Jaccard)
        N_tuple.append(sus)
    return N_tuple  #返回值没有第0行

def get_SFL_rank(file_path, test_dir_path, language):
    '''
    计算SBFL的怀疑度和排名
    返回值中：
    SFL_rank是排好序的语句怀疑度列表
    SFL_sus是排好序的语句怀疑度值
    N_tuple是源代码顺序的语句怀疑度值
    '''
    if language == 'py':
        passed_test_num, failed_test_num, lines_passed,  lines_failed = Coverage.get_python_cov_info(file_path, test_dir_path)
    elif language == 'cpp' or 'c':
        passed_test_num, failed_test_num, lines_passed,  lines_failed = Coverage.get_cpp_cov_info(file_path, test_dir_path)
    else:
        return [], [], []
    # print(passed_test_num, failed_test_num)
    # print(lines_passed,  lines_failed)
    N_tuple = cal_N_tuple(passed_test_num, failed_test_num, lines_passed,  lines_failed)
    # print(N_tuple)
    N_tuple_c = []
    for i in range(len(N_tuple)):
        N_tuple_c.append({
            'no': i + 1,
            'similarity': N_tuple[i]
        })
    N_tuple_c.sort(key=lambda s:(s['similarity']), reverse=True)
    SFL_rank = []
    SFL_sus = []
    for i in N_tuple_c:
        SFL_rank.append(i['no'])
        SFL_sus.append(i['similarity'])
    # print(SFL_sus)
    return SFL_rank, SFL_sus, N_tuple

def get_simple_coefficient(file_path, N_tuple, VSBFL_suspicion, language):
    '''
    这里是用于解决tie问题的一个小补丁：
    当SBFL所有句子怀疑度一样的时候，代表分支并没有出错。
    当所有变量怀疑度都为0的时候，代表变量的计算过程没有出错
    那么这个时候是不是应该只能是输出错误呢
    '''
    # print(file_path, N_tuple, VSBFL_suspicion)
    simple_coefficient = list(np.zeros(len(N_tuple)))
    flag = 0
    maxn = 0
    minn = 99
    for i in N_tuple:
        if i != 0:
            maxn = max(maxn, i)
            minn = min(minn, i)
    if maxn - minn > 0.001:
        flag = 1
    # print(maxn, minn)
    for var in VSBFL_suspicion:
        if VSBFL_suspicion[var] > 0.0001: 
            flag = 1
            break
    if flag == 1:
        simple_coefficient = N_tuple
    else:
        lines = util.read_file(file_path)
        for i, line in enumerate(lines):
            flag = False
            if language == 'cpp' or language == 'c':
                flag = (util.find_pos('cin', line) or util.find_pos('printf', line))
            elif language == 'py':
                flag = util.find_pos('print', line)
            # print(line)
            # print(util.find_pos('printf', line))
            if flag == True:
                simple_coefficient[i] = 1

    SFL_rank = []
    SFL_sus = []
    for i,num in enumerate(simple_coefficient):
        SFL_rank.append({
            'no': i + 1,
            'similarity': simple_coefficient[i]
        })
        SFL_sus.append(simple_coefficient[i])
    SFL_rank.sort(key=lambda s:(s['similarity']), reverse=True)
    SFL_sus.sort(reverse=True)
    return SFL_rank, SFL_sus, simple_coefficient

def cal_final_rank(VSBFL_rank, SFL_rank, N_tuple, variable_info):
    '''
    计算最终怀疑都列表
    '''
    final_rank = []

    # 新的排序方法
    # coefficient_list = []
    # final_rank_t = []
    # VSBFL_dic = {}
    # for item in VSBFL_rank:
    #     VSBFL_dic[item['name']] = item['value']
    # for i in range(len(variable_info)):
    #     coefficient = 0
    #     for variable in variable_info[i]:
    #         coefficient += VSBFL_dic[variable]
    #     if len(variable_info[i]) != 0:
    #         coefficient = 1 + coefficient / len(variable_info[i])
    #     else:
    #         coefficient = 1.0
    #     coefficient_list.append(coefficient)
    # for i in range(len(N_tuple)):
    #     final_rank_t.append({
    #         'no': i + 1,
    #         'pos': (1 + N_tuple[i]) * coefficient_list[i]
    #     })
    # # print(final_rank_t)
    # final_rank_t.sort(key=lambda s: s['pos'], reverse=True)
    # tmp_list = []
    # for i in range(len(final_rank_t)):
    #     if i != 0 and abs(final_rank_t[i]['pos'] - final_rank_t[i-1]['pos']) > 0.0001:
    #         final_rank.append(tmp_list)
    #         tmp_list = []
    #     tmp_list.append(final_rank_t[i]['no'])
    # if len(tmp_list) != 0:
    #     final_rank.append(tmp_list)

    # 新的排序方法2
    coefficient_list = []
    final_rank_t = []
    VSBFL_dic = {}
    for item in VSBFL_rank:
        VSBFL_dic[item['name']] = item['value']
    # print(N_tuple)
    for i in range(len(variable_info)):
        coefficient = 1.0
        for variable in variable_info[i]:
            coefficient += VSBFL_dic[variable]
        # if len(variable_info[i]) != 0:
        #     coefficient = 1 + coefficient / len(variable_info[i])
        # else:
        #     coefficient = 1.0
        coefficient_list.append(coefficient)
    print(coefficient_list)
    for i in range(len(N_tuple)):
        final_rank_t.append({
            'no': i + 1,
            'pos': (1.0 + N_tuple[i]) * coefficient_list[i]
        })
    print(final_rank_t)
    final_rank_t.sort(key=lambda s: s['pos'], reverse=True)
    tmp_list = []
    for i in range(len(final_rank_t)):
        if i != 0 and abs(final_rank_t[i]['pos'] - final_rank_t[i-1]['pos']) > 0.0001:
            final_rank.append(tmp_list)
            tmp_list = []
        tmp_list.append(final_rank_t[i]['no'])
    if len(tmp_list) != 0:
        final_rank.append(tmp_list)

    # 旧的排序方法
    # for item in VSBFL_rank:
    #     variable = item['name']
    #     cover_line_c = []
    #     for i in range(len(variable_info)):
    #         if variable in variable_info[i]:
    #             cover_line_c.append({
    #                 'no': i+1,
    #                 'pos': SFL_rank[i]
    #             })
    #     cover_line_c.sort(key=lambda s:(s['pos']))
    #     for i in cover_line_c:
    #         try:
    #             final_rank.index(i['no'])
    #         except:
    #             final_rank.append(i['no'])

    # print(final_rank)
    return final_rank

def cal_final_rank2(SFL_rank, SFL_sus):
    '''
    获取SBFL排名
    '''
    # print(SFL_rank, SFL_sus)
    final_rank = []
    tmp_list = []
    for i,num in enumerate(SFL_sus):
        if i == 0:
            tmp_list.append(SFL_rank[i])
            continue
        if num != SFL_sus[i-1]:
            final_rank.append(tmp_list)
            tmp_list = []
        tmp_list.append(SFL_rank[i])
    if len(tmp_list) > 0:
        final_rank.append(tmp_list)
    return final_rank,[]

def run_file(file_path, ac_file, test_dir_path, language):
    '''
    计算程序的最后怀疑度列表
    '''
    # print(file_path, ac_file)
    SFL_rank, SFL_sus, N_tuple = get_SFL_rank(file_path, test_dir_path, language)
    print(N_tuple)
    # final_rank, VSBFL_rank = VFL.get_VFL_list(file_path, N_tuple)
    final_rank, VSBFL_rank = cal_final_rank2(SFL_rank, SFL_sus)
    # VSBFL_suspicion, VSBFL_rank = vs.cal_VSBFL_rank(file_path, ac_file, test_dir_path, language)
    # SFL_rank, SFL_sus, N_tuple = get_simple_coefficient(file_path, N_tuple, VSBFL_suspicion, language)
    # # print(N_tuple)
    # variable_list = list(VSBFL_suspicion.keys())
    # variable_info = util.collect_variable_info(variable_list, file_path)
    # # print(variable_info)
    # final_rank = cal_final_rank(VSBFL_rank, SFL_rank, N_tuple, variable_info)
    print(final_rank)
    return final_rank, VSBFL_rank, N_tuple

def run_dir(file_dir_path, pair_info, test_dir_path):
    '''
    计算某个文件夹内所有文件的最后怀疑度列表
    '''
    file_list = os.listdir(file_dir_path)
    wb = openpyxl.load_workbook(res_file)
    if sys.platform == "linux":
        problem_id = file_dir_path.split('/')[-2]
    else:
        problem_id = file_dir_path.split('\\')[-2]
    wb.create_sheet(problem_id)
    ws = wb[problem_id]
    ws.append({'a':'name', 'b':'suspicion_rank', 'c':'suspicion'})
    for file in file_list:
        file_type = file.split('.')[-1]
        if file_type == 'c' or file_type == 'cpp' or file_type == 'py':
            wa_file_path = os.path.join(file_dir_path, file)
            ac_file_path = os.path.join(data_path, str(problem_id), 'AC_'+file_type, pair_info[file])
            print(wa_file_path, ac_file_path)
            try:
                final_rank, VSBFL_rank, N_tuple = run_file(wa_file_path, ac_file_path, test_dir_path, file_type)
                ws.append({'a':file, 'b':str(final_rank), 'c':str(N_tuple)})
                # util.add_file(res_file, file + '    ' + str(final_rank) + '    ' + str(VSBFL_rank) + '\n')
            except Exception:
                print(Exception)
                ws.append({'a':file, 'b':'[]', 'c':'[]'})
                # util.add_file(res_file, file + '    ' + 'contains error\n')
        # break
    # wb.save(res_file)
    return

def cal_time(file_dir_path, pair_info, test_dir_path):
    '''
    计算时间
    '''
    file_list = os.listdir(file_dir_path)
    wb = openpyxl.load_workbook(res_file)
    if sys.platform == "linux":
        problem_id = file_dir_path.split('/')[-2]
    else:
        problem_id = file_dir_path.split('\\')[-2]
    wb.create_sheet(problem_id)
    ws = wb[problem_id]
    ws.append({'a': 'solution_id', 'b': 'time'})
    for file in file_list:
        file_type = file.split('.')[-1]
        if file_type == 'c' or file_type == 'cpp' or file_type == 'py':
            wa_file_path = os.path.join(file_dir_path, file)
            ac_file_path = os.path.join(data_path, str(problem_id), 'AC_'+file_type, pair_info[file])
            # print(wa_file_path, ac_file_path)
            time_start=time.time()
            try:
                final_rank, VSBFL_rank = run_file(wa_file_path, ac_file_path, test_dir_path, file_type)
                # util.add_file(res_file, file + '    ' + str(final_rank) + '    ' + str(VSBFL_rank) + '\n')
            except Exception:
                print(Exception)
                # util.add_file(res_file, file + '    ' + 'contains error\n')
            time_end=time.time()
            ws.append({'a': file, 'b': str(time_end - time_start)})
        # break
    wb.save(res_file)

if __name__ == "__main__":

    # pair_info = find_pair_by_res(r'D:\fault_loc\VSFL-TCG\result\cluster.xlsx')
    # pair_info = find_pair_by_tag(os.path.join(data_path, '2810', 'Tag_c'))
    # for fm in ['Jaccard']:
    #     now_fm = fm
    #     for pro in ['270081_buggy.c']:
    #         file_path = os.path.join(data_path, '2810', 'WA_c', pro)
    #         # ac_file = os.path.join(data_path, '2810', 'AC_c', pair_info[pro])
    #         test_dir_path = os.path.join(data_path, '2810', 'TEST_DATA_TCG1')
    #         run_file(file_path, '', test_dir_path, 'c')

    for fm in ['Jaccard', 'Tarantula', 'Dstar', 'Ochiai', 'Op2']:
    # for fm in ['Jaccard']:
        now_fm = fm
        res_file = res_file_temp % (fm)
        problem_list = [2810, 2811, 2812, 2813, 2824, 2825, 2827, 2828, 2830, 2831, 2832, 2833, 2864, 2865, 2866, 2867, 2868, 2869, 2870, 2871]
        
        for problem in problem_list:
            # problem = 2810
            print(problem)
            pair_info = find_pair_by_tag(os.path.join(data_path, str(problem), 'Tag_c'))
            # pair_info = find_pair_by_res(r'D:\fault_loc\VSFL-TCG\result\cluster_op2.xlsx')
            dir_path = os.path.join(data_path, str(problem), 'WA_c')
            test_dir_path = os.path.join(data_path, str(problem), 'TEST_DATA_TCG1')
            run_dir(dir_path, pair_info, test_dir_path)
            # cal_time(dir_path, pair_info, test_dir_path)
            # break