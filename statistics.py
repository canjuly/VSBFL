# -*- coding: utf-8 -*-
import openpyxl
import os
import util
import numpy as np

def get_tag_info(tag_dir):
    '''
    获取真实的错误信息
    '''
    tag_info = {}
    file_list = os.listdir(tag_dir)
    for i in file_list:
        file_path = os.path.join(tag_dir, i)
        lines = util.read_file(file_path)
        name = lines[0].replace('\n', '')
        arr = lines[2].split(',')
        tag_info[name] = []
        for num in arr:
            tag_info[name].append(int(num))
    return tag_info


def cal_top_N_multi(rows, tag_info):
    '''
    统计Top-N的值(多错误)
    '''
    top_1 = 0
    top_3 = 0
    top_5 = 0
    top_10 = 0
    cnt = 0
    for i, row in enumerate(rows):
        # print(row[0].value, end='\t')
        # print(row[1].value, end='\t')
        # print(row[2].value)
        if i == 0:
            continue
        cnt += 1
        name = row[0].value
        final_rank = eval(row[1].value)
        ans = 0
        for j, column in enumerate(final_rank):
            ans += len(column)
            for num in tag_info[name]:
                if num in column:
                    if ans <= 1:
                        top_1 += 1
                    elif ans <= 3:
                        top_3 += 1
                    elif ans <= 5:
                        top_5 += 1
                    elif ans <= 10:
                        top_10 += 1
        # break
    # return top_1/cnt, top_3/cnt, top_5/cnt, top_10/cnt
    return top_1, top_3, top_5, top_10

def cal_top_N_first(rows, tag_info):
    '''
    统计Top-N的值(第一个错误)
    '''
    top_1 = 0
    top_3 = 0
    top_5 = 0
    top_10 = 0
    cnt = 0
    for i, row in enumerate(rows):
        if i == 0:
            continue
        # print(row[0].value, end='\t')
        # print(row[1].value, end='\t')
        # print(row[2].value)
        cnt += 1
        name = row[0].value
        final_rank = eval(row[1].value)
        ans = 0
        for j, column in enumerate(final_rank):
            flag = 0
            for num in tag_info[name]:
                ans += len(column)
                if num in column:
                    flag = 1
                    break
            if flag == 1:
                break
        if ans <= 1:
            top_1 += 1
        elif ans <= 3:
            top_3 += 1
        elif ans <= 5:
            top_5 += 1
        elif ans <= 10:
            top_10 += 1
        # break
    # return top_1/cnt, top_3/cnt, top_5/cnt, top_10/cnt
    return top_1, top_3, top_5, top_10

def cal_exam_multi(rows, tag_info):
    '''
    计算exam值(多错误)
    '''
    res = []
    for i, row in enumerate(rows):
        # print(row[0].value, end='\t')
        # print(row[1].value, end='\t')
        # print(row[2].value)
        if i == 0:
            continue
        ans = 0
        name = row[0].value
        final_rank = eval(row[1].value)
        for num in tag_info[name]:
            no = 0
            for j, column in enumerate(final_rank):
                no += len(column)
                if num in column:
                    ans += no
                    break
        ans = ans / len(tag_info[name])
        res.append(ans)
    return res

def cal_exam_first(rows, tag_info):
    '''
    计算exam值(第一个错误)
    '''
    res = []
    for i, row in enumerate(rows):
        if i == 0:
            continue
        # print(row[0].value, end='\t')
        # print(row[1].value, end='\t')
        # print(row[2].value)
        name = row[0].value
        final_rank = eval(row[1].value)
        ans = 0
        for j, column in enumerate(final_rank):
            ans += len(column)
            flag = 0
            for num in tag_info[name]:
                if num in column:
                    flag = 1
                    break
            if flag == 1:
                break
        # if flag == 0:
        #     ans = 
        res.append(ans)
    return res

def get_code_exam(sheet, tag_info):
    '''
    '''
    sum = 0
    for i, row in enumerate(sheet.rows):
        if i == 0:
            continue
        name = row[0].value
        final_rank = eval(row[1].value)
        res1 = 0
        ans1 = 0
        res2 = 0
        ans2 = 1
        flag = 1
        for j, column in enumerate(final_rank):
            ans1 += len(column)
            for num in tag_info[name]:
                if num in column:
                    # if flag == 1:
                        # res1 += ans1
                        # res2 += ans2
                        # flag = 0
                    res1 += ans1
                    res2 += ans2
                    flag = 0
            ans2 += len(column)
        if flag == 1:
            res1 = ans1
            res2 = ans1
        else:
            res1 = res1 / len(tag_info[name])
            res2 = res2 / len(tag_info[name])
        if ans1 != 0:
            sheet.cell(i + 1, 4).value = (res1 + res2) / 2 / ans1
            # sheet.cell(i + 1, 5).value = (res1 + res2) / 2 / ans1
            # sheet.cell(i + 1, 6).value = ''
        else:
            sheet.cell(i + 1, 4).value = 1
            # sheet.cell(i + 1, 5).value = 1
            # sheet.cell(i + 1, 6).value = ''
        sum += (res1 + res2) / 2 / ans1
    return sum

def statistical_fl_results(file_path, tag_root_dir):
    '''
    统计错误定位的实验结果
    '''
    if file_path.split('.')[-1] != 'xlsx' and file_path.split('.')[-1] != 'xls':
        print('not an execl file')
        return
    wb = openpyxl.load_workbook(file_path) # 读取 工作簿
    ws_first = wb.worksheets[0]  
    ws_first.title = 'first'
    ws_multi = wb.worksheets[1]
    ws_multi.title = 'multi'
    ws_first.append({'a':'problem_id', 'b':'exam', 'c':'top_1', 'd':'top_3', 'e':'top_5', 'f': 'top_10','g': 'total_num'})
    ws_multi.append({'a':'problem_id', 'b':'exam', 'c':'top_1', 'd':'top_3', 'e':'top_5', 'f': 'top_10','g': 'total_num'})
    ans1 = 0
    ans2 = 0
    for i, sheet in enumerate(wb):
        if i == 0 or i == 1:
            continue
        # if sheet.title != '2812':
        #     continue
        print(sheet.title)
        # row_max = sheet.max_row # 获取最大行
        # con_max = sheet.max_column # 获取最大列
        tag_dir = os.path.join(tag_root_dir, sheet.title, 'Tag_c')
        tag_info = get_tag_info(tag_dir)
        # for i in tag_info:
        #     ans1 += len(tag_info[i])
        # continue
        ans1 += get_code_exam(sheet, tag_info) # 给每道题添加exam
        ans2 += (sheet.max_row - 1)
        exam = cal_exam_first(sheet.rows, tag_info)
        top_1, top_3, top_5, top_10 = cal_top_N_first(sheet.rows, tag_info)
        ws_first.append({'a':sheet.title, 'b':np.mean(exam), 'c':top_1, 'd':top_3, 'e':top_5, 'f': top_10, 'g':sheet.max_row - 1})
        
        top_1, top_3, top_5, top_10 = cal_top_N_multi(sheet.rows, tag_info)
        exam = cal_exam_multi(sheet.rows, tag_info)
        ws_multi.append({'a':sheet.title, 'b':np.mean(exam), 'c':top_1, 'd':top_3, 'e':top_5, 'f': top_10, 'g':sheet.max_row - 1})
        
        # print(exam)
        # print(top_1, top_3, top_5, top_10)
        # break
    print(ans1)
    wb.save(file_path)
    return

def cal_exact_result(rows):
    '''
    计算匹配到的代码包含数据集配对代码的数量
    '''
    name = ''
    cnt = 0   #代码份数
    ans1 = 0  #统计结果包含配对代码的
    ans2 = 0  #统计结果恰好是配对代码的
    for i, row in enumerate(rows):
        if i == 0:
            continue
        if row[0].value != None:
            cnt += 1
            name = row[0].value.split('_')[0]
        if row[1].value.find('[') >= 0:
            res_list = eval(row[1].value)
            res_list = list(map(lambda x: x.split('_')[0], res_list))
            # print(res_list)
            if name in res_list:
                ans1 += 1
            if len(res_list) == 1 and res_list[0] == name:
                ans2 += 1
    print(ans1, ans2, cnt)
    return ans1, ans2, ans1/cnt, ans2/cnt

def statistical_parse_results(file_path):
    '''
    统计匹配正确代码的结果
    '''
    if file_path.split('.')[-1] != 'xlsx' and file_path.split('.')[-1] != 'xls':
        print('not an execl file')
        return
    wb = openpyxl.load_workbook(file_path)
    ws = wb.worksheets[0]
    ws.append({'a':'name', 'b':'contain_cnt', 'c':'contain_ratio', 'd':'exact_cnt', 'e':'exact_ratio'})
    for i, sheet in enumerate(wb):
        if i == 0:
            continue
        print(sheet.title)
        ans1, ans2, ratio1, ratio2 = cal_exact_result(sheet.rows)
        ws.append({'a':sheet.title, 'b':ans1, 'c':ratio1, 'd': ans2, 'e': ratio2})
        # break
    wb.save(file_path)
    return

def statistical_type_result(file_path, tag_root_dir):
    '''
    '''
    if file_path.split('.')[-1] != 'xlsx' and file_path.split('.')[-1] != 'xls':
        print('not an execl file')
        return
    wb = openpyxl.load_workbook(file_path) # 读取 工作簿
    ws = wb.worksheets[0]
    ws.append({'a': 'type', 'b': 'count', 'c': 'first', 'd': 'multi'})
    simple_expression = {
        'count': 0, 'first': 0, 'multi': 0,
    }
    conditions = {
        'first': 0, 'multi': 0, 'count': 0,
    }
    loops = {
        'first': 0, 'multi': 0, 'count': 0,
    }
    for i, sheet in enumerate(wb):
        if i == 0:
            continue
        print(sheet.title)
        for j, row in enumerate(sheet.rows):
            if j == 0:
                continue
            if sheet.title in ['2810', '2811', '2812', '2813']:
                simple_expression['first'] += row[4].value
                simple_expression['multi'] += row[3].value
                simple_expression['count'] += 1
            elif sheet.title in ['2824', '2825', '2827', '2828', '2830', '2831', '2832', '2833']:
                conditions['first'] += row[4].value
                conditions['multi'] += row[3].value
                conditions['count'] += 1
            elif sheet.title in ['2864', '2865', '2866', '2867', '2868', '2869', '2870', '2871']:
                loops['first'] += row[4].value
                loops['multi'] += row[3].value
                loops['count'] += 1
    print(simple_expression, conditions, loops)
    ws.append({'a': 'simple_expression', 'b': simple_expression['count'], 'c': simple_expression['first'] / simple_expression['count'], 'd': simple_expression['multi'] / simple_expression['count']})
    ws.append({'a': 'conditions', 'b': conditions['count'], 'c': conditions['first'] / conditions['count'], 'd': conditions['multi'] / conditions['count']})
    ws.append({'a': 'loops', 'b': loops['count'], 'c': loops['first'] / loops['count'], 'd': loops['multi'] / loops['count']})
    wb.save(file_path)
    return

if __name__ == "__main__":
    
    # for fm in ['Jaccard', 'Dstar', 'Ochiai', 'Op2', 'Tarantula']:
    for fm in ['Jaccard']:
        # for fc in ['sbfl', 'mbfl', 'op2', 'vfl', 'res_op2']:
        for fc in ['op2']:
            print(fm, fc)
            file_path = r'C:\Users\ShenJitao\Desktop\cppsnooper\instrumentation\result\sbfl(%s).xlsx' % (fm)
            tag_dir = r'D:\fault_loc\ITSP-data'
            statistical_fl_results(file_path, tag_dir)
            # statistical_type_result(file_path, tag_dir)
            # break

    # file_path = r'E:\fault_loc\VSFL-TCG\result\cluster_op2.xlsx'
    # statistical_parse_results(file_path)